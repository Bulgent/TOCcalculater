import csv
import glob
from statistics import stdev, mean
import numpy as np
import pandas as pd
import collections
import copy
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side


# 項目を含む配列数を検索
def search_row_num(lists, content_name):
    ret = None
    for i, lst in enumerate(lists):
        if content_name in lst:
            ret = i
            break
    return ret
# Ave.Area抽出
def extract_AveArea(lists):
    row_AveArea = search_row_num(lists, 'Ave. Area')
    index_AveArea = lists[row_AveArea].index('Ave. Area')
    return float(lists[row_AveArea+1][index_AveArea])

def extract_STDConc(lists):
    row_STDConc = search_row_num(lists, "STD Conc")
    index_STDConc = lists[row_STDConc].index("STD Conc")
    return float(lists[row_STDConc+1][index_STDConc])

def extract_SampleName(lists):
    row_SampleName = search_row_num(lists, "Sample Name")
    index_SampleName = lists[row_SampleName].index("Sample Name")
    return lists[row_SampleName+1][index_SampleName].replace(" ","")

def extract_VialNo(lists):
    row_VialNo = search_row_num(lists, "VialNo")
    index_VialNo = lists[row_VialNo].index("VialNo")
    return int(lists[row_VialNo+1][index_VialNo])

def checkAreaBalance(lists):
    row_Area = search_row_num(lists, "Area")
    index_Area = lists[row_Area].index("Area")
    areas = []
    for i in range(row_Area+1, len(lists)):
        areas.append(lists[i][index_Area])
    coefficient_variation = stdev(areas)/mean(areas)
    return coefficient_variation

def makeLine(lst_x,lst_y):
    arr_x = np.array(lst_x)
    arr_y = np.array(lst_y)
    n = len(arr_x)
    mean_x = sum(arr_x)/n
    mean_y = sum(arr_y)/n
    t_xx = sum((arr_x-mean_x)**2)
    t_yy = sum((arr_y-mean_y)**2)
    t_xy = sum((arr_x-mean_x)*(arr_y-mean_y))
    slope = t_xy/t_xx
    intercept = (1/n)*sum(arr_y)-(1/n)*slope*sum(arr_x)
    predict_x = intercept+slope*arr_x
    residual_y = arr_y-predict_x
    r2 = 1-(sum(residual_y**2))/t_yy
    return slope, intercept, r2

def makeStandard(datum, STD_vial_no_lst, blank_vial_no_lst):

    ave_area_lst = []
    std_conc_lst = []
    vial_no_lst = []

    for STD_vial_no in STD_vial_no_lst:
        lst = datum[STD_vial_no]
        vial_no_lst.append(STD_vial_no)
        ave_area_lst.append(extract_AveArea(lst))
        std_conc_lst.append(extract_STDConc(lst))

    #複数のブランクから最も優れたブランクを利用
    r2_max = 0
    for blank_vial_no in blank_vial_no_lst:
        ave_area_lst_copied = copy.copy(ave_area_lst)
        std_conc_lst_copied = copy.copy(std_conc_lst)
        vial_no_lst_copied = copy.copy(vial_no_lst)
        lst = datum[blank_vial_no]
        ave_area_lst_copied.insert(0, extract_AveArea(lst))
        std_conc_lst_copied.insert(0, 0)
        vial_no_lst_copied.insert(0, blank_vial_no)

        _, _, r2 = makeLine(ave_area_lst_copied, std_conc_lst_copied)
        
        if r2 > r2_max:
            df_ret = pd.DataFrame({"Ave.Area":ave_area_lst_copied,"Conc.":std_conc_lst_copied, "VialNo":vial_no_lst_copied})
            r2_max = r2
            ret_blank_vial_no = blank_vial_no

    return df_ret, ret_blank_vial_no

def getSeveralData(lists):
    vial_no = extract_VialNo(lists)
    sample_name = extract_SampleName(lists)
    ave_area = extract_AveArea(lists)
    return vial_no, sample_name, ave_area

def checkSampleCondition(lists):
    '''
    csvを読み取り結合したデータに対して適用
    以下のルールに従い {VialNo : サンプルの状態} 
    blank: メジャーなサンプル名以外のサンプル
    sample: メジャーなサンプル名
    standard: TC, ICの文字列が含まれているサンプル
    '''
    ret = {} # {VialNo : Sample Name}
    for num, lst in enumerate(lists):
        if "Date" in lst:
            index_sampleName = lst.index("Sample Name")
            index_vialNo = lst.index("VialNo")
            sampleName = lists[num+1][index_sampleName].replace(" ","")
            vialNo     = int(lists[num+1][index_vialNo].replace(" ",""))
            if "TC" in sampleName:
                label = "standard_TC"
            elif "IC" in sampleName:
                label = "standard_IC"
            else:
                label = "sample_" + sampleName[-1]
            
            ret[vialNo] = label
    content_count = collections.Counter(list(ret.values()))
    iter = 0
    while True:
        sample_label = content_count.most_common()[iter][0]
        if "TC" in sample_label or "IC" in sample_label:
            iter += 1
            continue
        else:
            break

    for k in ret.keys():
        if "TC" not in ret[k] and "IC" not in ret[k] and sample_label != ret[k]:
            ret[k] = "blank"
        elif sample_label == ret[k]:
            ret[k] = "sample"

    return ret

def getVialNo(dict, sampleName):
    '''
    checkSampleConditionの出力値を持ってくる
    '''
    ret = []
    for k, v in dict.items():
        if v == sampleName:
            ret.append(k)
    return ret

def makeSTDSheet(book,sheet_name):
    # シートを取得
    sheet = book[sheet_name]
    # シートに検量線を記述
    sheet.cell(row=1,column=5).value = 'SLOPE'
    sheet.cell(row=1,column=6).value = '=SLOPE(B2:B4,A2:A4)'
    sheet.cell(row=2,column=5).value = 'INTERCEPT'
    sheet.cell(row=2,column=6).value = '=INTERCEPT(B2:B4,A2:A4)'
    sheet.cell(row=3,column=5).value = 'R2'
    sheet.cell(row=3,column=6).value = '=RSQ(B2:B4,A2:A4)'

    # 背景色, 太字設定
    for row_num in range(1,4):
        for column_num in range(5, 6):
            sheet.cell(row=row_num,column=column_num).font = Font(bold=True)
            sheet.cell(row=row_num,column=column_num).fill = PatternFill(patternType="solid", fgColor="e5d0e0")

    # 背景色, 太字設定
    for row_num in range(1,2):
        for column_num in range(1, 4):
            sheet.cell(row=row_num,column=column_num).font = Font(bold=True)
            sheet.cell(row=row_num,column=column_num).fill = PatternFill(patternType="solid", fgColor="c8deee")

    # 罫線の作成
    side1 = Side(style='thin', color='000000')
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row_num in range(1,5):
        for column_num in range(1, 4):
            sheet.cell(row = row_num, column=column_num).border = border_aro
    for row_num in range(1,4):
        for column_num in range(5, 7):
            sheet.cell(row = row_num, column=column_num).border = border_aro

    # グラフを記述
    chart = ScatterChart()
    xvalues = Reference(sheet, min_col=1, min_row=2, max_row=4)
    values = Reference(sheet, min_col=2, min_row=1, max_row=4)
    series = Series(values, xvalues, title_from_data=True)
    series.graphicalProperties.line.noFill=True
    chart.series.append(series)
    chart.series[0].marker.symbol = "circle"
    chart.series[0].marker.size = 10
    series.marker.spPr.solidFill = "8996a3" 
    tl = Trendline(trendlineType='linear')
    chart.ser[-1].trendline = tl
    sheet.add_chart(chart,"A6")

# DataReportのフォルダパスを指定
path_folder = r".\Results"
results = glob.glob(path_folder +r"\Result*.csv")

# csvからデータを抽出
readers = []
for result in results:
    with open(result, encoding='ANSI', newline='') as f:
        reader = [i for i in csv.reader(f) if len(i)>=6] #F列以降もある行を抽出
        readers += reader

sample_condition = checkSampleCondition(readers)

# TC, IC毎にデータを纏める
# {vialNo: [["Date","Sample No",..],["1/12/2022  11:46:20", "1",..],["Injection #","Start ID"],..],vialNo: [],...}
# サンプル単位でTC, IC別に配列をまとめる
datum_TC = {}
datum_IC = {}
temp_row = 0
for num_row, data in enumerate(readers):
    if "Date" in data and num_row != 0:
        vial_no = extract_VialNo(readers[temp_row:num_row])
        if "TC" in readers[temp_row+1]:
            datum_TC[vial_no] = readers[temp_row:num_row]
        if "IC" in readers[temp_row+1]:
            datum_IC[vial_no] = readers[temp_row:num_row]
        temp_row = num_row
#最後のブロックについても同様に処理
vial_no = extract_VialNo(readers[temp_row:len(readers)])
if "TC" in readers[temp_row+1]:
    datum_TC[vial_no] = readers[temp_row:len(readers)]
if "IC" in readers[temp_row+1]:
    datum_IC[vial_no] = readers[temp_row:len(readers)]

# 検量線作成(ブランクは最後のバイアルを使用)
df_standard_TC, blank_vial_no = makeStandard(datum_TC, getVialNo(sample_condition, "standard_TC"), getVialNo(sample_condition, "blank"))
df_standard_IC, blank_vial_no = makeStandard(datum_IC, getVialNo(sample_condition, "standard_IC"), getVialNo(sample_condition, "blank"))

# サンプルデータの整形
# [{"VialNo":,"Label":,"TC_Ave.Area":,"IC_Ave.Area":},]
display_lst = []
display_vial_no_lst = getVialNo(sample_condition, "blank") + getVialNo(sample_condition, "sample")
display_vial_no_lst.sort()
for vial_no in display_vial_no_lst:
    temp_dict = {}
    lst_tc = datum_TC[vial_no]
    lst_ic = datum_IC[vial_no]
    temp_dict["VialNo"] = vial_no
    temp_dict["Label"] = sample_condition[vial_no]
    temp_dict["TC_Ave.Area"] = extract_AveArea(lst_tc)
    temp_dict["IC_Ave.Area"] = extract_AveArea(lst_ic)
    display_lst.append(temp_dict)

df = pd.DataFrame(display_lst)

#pandasでの処理結果をExcelに記述
output_excel_path = path_folder+r"/toc_result.xlsx"
with pd.ExcelWriter(output_excel_path) as writer:
    df.to_excel(writer, sheet_name="data", index=False)
    df_standard_TC.to_excel(writer, sheet_name="TC_standard",index=False)
    df_standard_IC.to_excel(writer, sheet_name="IC_standard",index=False)

# openpyxlでExcelを編集
book = openpyxl.load_workbook(output_excel_path)

# 検量線シートを編集
makeSTDSheet(book,'TC_standard')
makeSTDSheet(book,'IC_standard')

# TOC結果シートを編集
sheet_data = book['data']
sheet_data.cell(row=1,column=5).value = 'TC.Conc(ppm)'
sheet_data.cell(row=1,column=6).value = 'IC.Conc(ppm)'
sheet_data.cell(row=1,column=7).value = 'TOC.Conc(ppm)'
# TC, ICの計算式を記述
for row_num in range(2,sheet_data.max_row+1):
    sheet_data.cell(row=row_num,column=5).value = f'=C{row_num}*TC_standard!F1+TC_standard!F2'
    sheet_data.cell(row=row_num,column=6).value = f'=D{row_num}*IC_standard!F1+IC_standard!F2'
    sheet_data.cell(row=row_num,column=7).value = f'=E{row_num} - F{row_num}'

# レイアウト
# 背景色を付ける
for column_num in range(1,8):
    sheet_data.cell(row=1, column=column_num).font = Font(bold=True)
    sheet_data.cell(row=1, column=column_num).fill = PatternFill(patternType="solid", fgColor="eee87a")

# 罫線作成
side1 = Side(style='thin', color='000000')
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row_num in range(1,sheet_data.max_row+1):
    for column_num in range(1, sheet_data.max_column+1):
        sheet_data.cell(row = row_num, column=column_num).border = border_aro

# 列幅の調整
num2alpha = lambda c: chr(c+64)
for col in sheet_data.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = (max_length + 2) * 1.2
    sheet_data.column_dimensions[num2alpha(column)].width = adjusted_width

# 保存する
book.save(output_excel_path)