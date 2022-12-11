import csv
import glob
from statistics import stdev, mean
import numpy as np
import pandas as pd
import openpyxl


# 項目を含む配列数を検索
def search_row_num(lists, name):
    ret = None
    for i, lst in enumerate(lists):
        if name in lst:
            ret = i
            break
    return ret

# Ave.Area抽出
def extract_AveArea(lists):
    row_AveArea = search_row_num(lists, 'Ave. Area')
    index_AveArea = lists[row_AveArea].index('Ave. Area')
    return float(lists[row_AveArea+1][index_AveArea])

def extract_STDConc(lists):
    row_STDConc = search_row_num(lists, "STD Conc")
    index_STDConc = lists[row_STDConc].index("STD Conc")
    return float(lists[row_STDConc+1][index_STDConc])

def extract_SampleName(lists):
    row_SampleName = search_row_num(lists, "Sample Name")
    index_SampleName = lists[row_SampleName].index("Sample Name")
    return lists[row_SampleName+1][index_SampleName].replace(" ","")

def extract_VialNo(lists):
    row_VialNo = search_row_num(lists, "VialNo")
    index_VialNo = lists[row_VialNo].index("VialNo")
    return int(lists[row_VialNo+1][index_VialNo])

def checkAreaBalance(lists):
    row_Area = search_row_num(lists, "Area")
    index_Area = lists[row_Area].index("Area")
    areas = []
    for i in range(row_Area+1, len(lists)):
        areas.append(lists[i][index_Area])
    coefficient_variation = stdev(areas)/mean(areas)
    return coefficient_variation

def makeLine(lst_x,lst_y):
    arr_x=np.array(lst_x)
    arr_y=np.array(lst_y)
    n=len(arr_x)
    mean_x = sum(arr_x)/n
    mean_y = sum(arr_y)/n
    t_xx = sum((arr_x-mean_x)**2)
    t_yy = sum((arr_y-mean_y)**2)
    t_xy = sum((arr_x-mean_x)*(arr_y-mean_y))
    slope = t_xy/t_xx
    intercept = (1/n)*sum(arr_y)-(1/n)*slope*sum(arr_x)
    predict_x = intercept+slope*arr_x
    residual_y = arr_y-predict_x
    r2 = 1-(sum(residual_y**2))/t_yy
    return slope, intercept, r2

def makeStandard(datum, vial_no_lst, std_conc_lst):
    # TC: vial_no_lst = (1,2,3)
    # IC: vial_no_lst = (1,4,5)
    ave_area_lst = []
    for lst in datum:
        vial_no = extract_VialNo(lst)
        #if vial_no == 1:
        #    ave_area_lst.append(extract_AveArea(lst))
        if vial_no in vial_no_lst:
            ave_area_lst.append(extract_AveArea(lst))
    ave_area_lst.sort()
    #print()
    print(ave_area_lst)
    df_ret = pd.DataFrame({"Ave.Area":ave_area_lst,"Conc.":std_conc_lst})
    
    slope, intercept, r2 = makeLine(ave_area_lst, std_conc_lst)
    #df_ret = pd.DataFrame({"Ave.Area":ave_area_lst,"Conc.":std_conc_lst})
    
    return slope, intercept, r2, df_ret

def getSeveralData(lists):
    vial_no = extract_VialNo(lists)
    sample_name = extract_SampleName(lists)
    ave_area = extract_AveArea(lists)
    return vial_no, sample_name, ave_area

# DataReportのフォルダパスを指定
path_folder = r".\test"
results = glob.glob(path_folder +r"\Result*.csv")

# csvからデータを抽出
readers = []
for result in results:
    with open(result, encoding='ANSI', newline='') as f:
        reader = [i for i in csv.reader(f) if len(i)>=6] #F列以降もある行を抽出
        readers += reader

# TC, IC毎にデータを纏める
# [[["Date","Sample No",..],["1/12/2022  11:46:20", "1",..],["Injection #","Start ID"],..],[],...]
datum_TC = []
datum_IC = []
temp_row = 0
for num_row, data in enumerate(readers):
    # サンプル単位でTC, IC別に配列をまとめる
    if "Date" in data and num_row != 0:
        if "TC" in readers[temp_row+1]:
            datum_TC.append(readers[temp_row:num_row])
        if "IC" in readers[temp_row+1]:
            datum_IC.append(readers[temp_row:num_row])
        temp_row = num_row
#最後のブロックについても同様に処理
if "TC" in readers[temp_row+1]:
    datum_TC.append(readers[temp_row:len(readers)])
if "IC" in readers[temp_row+1]:
    datum_IC.append(readers[temp_row:len(readers)])

# 検量線作成(ブランクは最後のバイアルを使用)
TC_slope, TC_intercept, TC_r2, df_standard_TC = makeStandard(datum_TC, (len(datum_TC)+2,2,3), (0,10,50))
IC_slope, IC_intercept, IC_r2, df_standard_IC = makeStandard(datum_IC, (len(datum_TC)+2,4,5), (0,10,20))

# サンプルデータの整形
# [{"VialNo":,"TC_Ave.Area":,"IC_Ave.Area":,"TC_Conc":,"IC_Conc":},]
sample_datum = [None]*(len(datum_IC) - 4) #サンプルのデータ配列作成(検量線は除く)
for data_TC, data_IC in zip(datum_TC, datum_IC):
    vial_no_tc = extract_VialNo(data_TC)
    vial_no_ic = extract_VialNo(data_IC)
    sample_name_tc = extract_SampleName(data_TC)
    sample_name_ic = extract_SampleName(data_IC)
    ave_area_tc = extract_AveArea(data_TC)
    ave_area_ic = extract_AveArea(data_IC)
    if sample_name_tc == "4": # サンプルを抽出
        if sample_datum[vial_no_tc -6]:
            temp_dict = sample_datum[vial_no_tc -6]
        else:
            temp_dict = {}
            temp_dict["VialNo"] = vial_no_tc
        temp_dict["TC_Ave.Area"] = ave_area_tc
        #temp_dict["TC_Conc"] = ave_area_tc * TC_slope + TC_intercept
        sample_datum[vial_no_tc -6] = temp_dict
    if sample_name_ic == "4": # サンプルを抽出
        if sample_datum[vial_no_ic -6]:
            temp_dict = sample_datum[vial_no_ic -6]
        else:
            temp_dict = {}
            temp_dict["VialNo"] = vial_no_ic
        temp_dict["IC_Ave.Area"] = ave_area_ic
        #temp_dict["IC_Conc"] = ave_area_ic * IC_slope + IC_intercept
        sample_datum[vial_no_ic -6] = temp_dict

df = pd.DataFrame(sample_datum)
#df['TOC'] = df['TC_Conc'] - df['IC_Conc']

print(df)
df.to_csv(path_folder+r"/toc_result.csv", index=False)

output_excel_path = path_folder+r"/toc_result.xlsx"
with pd.ExcelWriter(output_excel_path) as writer:
    df.to_excel(writer, sheet_name="data", index=False)
    df_standard_TC.to_excel(writer, sheet_name="TC_standard",index=False)
    df_standard_IC.to_excel(writer, sheet_name="IC_standard",index=False)

# openpyxlで処理
# ブックを取得
book = openpyxl.load_workbook(output_excel_path)

# シートを取得
sheet_TC_standard = book['TC_standard']
# シートを取得
sheet_TC_standard.cell(row=1,column=4).value = 'SLOPE'
sheet_TC_standard.cell(row=1,column=5).value = '=SLOPE(A2:A4,B2:B4)'
sheet_TC_standard.cell(row=2,column=4).value = 'INTERCEPT'
sheet_TC_standard.cell(row=2,column=5).value = '=INTERCEPT(A2:A4,B2:B4)'
sheet_TC_standard.cell(row=3,column=4).value = 'R2'
sheet_TC_standard.cell(row=3,column=5).value = '=RSQ(A2:A4,B2:B4)'

# シートを取得
sheet_IC_standard = book['IC_standard']
# シートを取得
sheet_IC_standard.cell(row=1,column=4).value = 'SLOPE'
sheet_IC_standard.cell(row=1,column=5).value = '=SLOPE(A2:A4,B2:B4)'
sheet_IC_standard.cell(row=2,column=4).value = 'INTERCEPT'
sheet_IC_standard.cell(row=2,column=5).value = '=INTERCEPT(A2:A4,B2:B4)'
sheet_IC_standard.cell(row=3,column=4).value = 'R2'
sheet_IC_standard.cell(row=3,column=5).value = '=RSQ(A2:A4,B2:B4)'

# 保存する
book.save(output_excel_path)